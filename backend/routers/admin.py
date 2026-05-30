from fastapi import APIRouter, Query, HTTPException, Depends, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta, timezone
import uuid

from db import get_db
from models import User, Appointment, Order, LabBooking, AuditLog, Medicine, LabTest, WarehouseInventory, RegistrationRequest
from auth_utils import require_role, hash_password
from utils.helpers import model_to_dict

router = APIRouter()

@router.get("/dashboard")
def get_dashboard(current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    patients_count = db.query(User).filter(User.role == "patient").count()
    doctors_count = db.query(User).filter(User.role == "doctor").count()
    pharmacies_count = db.query(User).filter(User.role == "pharmacy").count()
    labs_count = db.query(User).filter(User.role.in_(["lab", "radiology"])).count()
    warehouses_count = db.query(User).filter(User.role == "warehouse").count()
    total_users = db.query(User).count()
    total_apts = db.query(Appointment).count()
    total_orders = db.query(Order).count()
    total_lab_bookings = db.query(LabBooking).count()
    pending_verifications = db.query(User).filter(User.verified == False).count()
    recent_activity = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).limit(10).all()
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    daily_apts = db.query(
        func.substr(Appointment.created_at, 1, 10).label('date'),
        func.count(Appointment.id)
    ).filter(Appointment.created_at >= seven_days_ago).group_by('date').all()
    daily_apts_data = {d: c for d, c in daily_apts}
    return {
        "total_users": total_users, "patients": patients_count, "doctors": doctors_count,
        "pharmacies": pharmacies_count, "labs": labs_count, "warehouses": warehouses_count,
        "total_appointments": total_apts, "total_orders": total_orders,
        "total_lab_bookings": total_lab_bookings, "pending_verifications": pending_verifications,
        "recent_activity": [model_to_dict(a) for a in recent_activity],
        "daily_appointments": daily_apts_data,
    }

@router.get("/users")
def list_all_users(role: str = Query(None), current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    return [model_to_dict(u, ["password"]) for u in query.all()]

@router.post("/users")
def create_user(req: dict, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    existing_email = db.query(User).filter(User.email == req.get("email")).first()
    if existing_email:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم مسبقاً")
    if req.get("phone"):
        existing_phone = db.query(User).filter(User.phone == req.get("phone")).first()
        if existing_phone:
            raise HTTPException(status_code=400, detail="رقم الهاتف مستخدم مسبقاً")
    new_id = f"new_{uuid.uuid4().hex[:8]}"
    role = req.get("role", "patient")
    photo = ""
    if role == 'doctor':
        photo = "/assets/doctors_sample/doctor male 3.png"
    password = req.get("password") or "123456"
    new_user = User(
        id=new_id, role=role, name=req.get("name", "").strip(), email=req.get("email"),
        password=hash_password(password), phone=req.get("phone", ""), city=req.get("city", ""),
        photo=photo, specialization=req.get("specialization") if role == 'doctor' else None,
        is_active=True, verified=True, created_at=datetime.now(timezone.utc).isoformat()
    )
    db.add(new_user); db.commit(); db.refresh(new_user)
    return model_to_dict(new_user, ["password"])

@router.put("/users/{user_id}/verify")
def verify_user(user_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "المستخدم غير موجود")
    user.verified = True
    db.commit(); db.refresh(user)
    return model_to_dict(user, ["password"])

@router.put("/users/{user_id}/toggle-active")
def toggle_active(user_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "المستخدم غير موجود")
    user.is_active = not user.is_active
    db.commit(); db.refresh(user)
    return model_to_dict(user, ["password"])

@router.put("/users/{user_id}/toggle-featured")
def toggle_featured(user_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "المستخدم غير موجود")
    user.is_featured = not user.is_featured
    db.commit(); db.refresh(user)
    return model_to_dict(user, ["password"])

@router.delete("/users/{user_id}")
def delete_user(user_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        db.delete(user); db.commit()
    return {"message": "تم حذف المستخدم"}

@router.get("/audit-logs")
def get_audit_logs(current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).all()
    return [model_to_dict(l) for l in logs]

@router.get("/stats")
def get_stats(current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    apts_by_status_query = db.query(Appointment.status, func.count(Appointment.id)).group_by(Appointment.status).all()
    appointments_by_status = {status: count for status, count in apts_by_status_query}
    orders_by_status_query = db.query(Order.status, func.count(Order.id)).group_by(Order.status).all()
    orders_by_status = {status: count for status, count in orders_by_status_query}
    return {
        "appointments_by_status": appointments_by_status, "orders_by_status": orders_by_status,
        "medicines_count": db.query(Medicine).count(), "lab_tests_count": db.query(LabTest).count(),
        "warehouse_items_count": db.query(WarehouseInventory).count(),
    }

@router.get("/registration-requests")
def list_registration_requests(current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    requests = db.query(RegistrationRequest).filter(RegistrationRequest.status == "pending").all()
    return [{"id": r.id, "role": r.role, "email": r.email, "data": r.data, "created_at": r.created_at} for r in requests]

@router.post("/registration-requests/{request_id}/approve")
def approve_registration(request_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    req = db.query(RegistrationRequest).filter(RegistrationRequest.id == request_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    if req.status != "pending":
        raise HTTPException(status_code=400, detail="الطلب تمت معالجته مسبقاً")
    data = req.data
    new_user = User(
        id=f"{req.role[0]}_{uuid.uuid4().hex[:8]}",
        role=req.role, name=f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
        email=req.email, password=hash_password(data.get("password", "123456")),
        phone=data.get('phone', ''), city=data.get('city', ''), photo=data.get('photo', ''),
        address=data.get('address', ''),
        clinic_name=data.get('clinic_name') if req.role == 'doctor' else None,
        clinic_address=data.get('clinic_address') if req.role == 'doctor' else None,
        price_per_session=data.get('price_per_session') if req.role == 'doctor' else data.get('home_service_fee'),
        experience_years=data.get('experience_years') if req.role == 'doctor' else None,
        available_hours=data.get('available_hours') if req.role == 'doctor' else None,
        working_hours=data.get('working_hours') if req.role == 'doctor' else None,
        specialization=data.get('specialization') if req.role == 'doctor' else None,
        open_hours=data.get('open_hours') if req.role in ('pharmacy', 'lab', 'radiology', 'warehouse') else None,
        home_service_fee=float(data.get('home_service_fee', 0) or 0) if req.role in ('lab', 'radiology') else None,
        has_home_service=bool(data.get('has_home_service', False)) if req.role in ('lab', 'radiology') else False,
        services=data.get('services') if req.role in ('lab', 'radiology') else None,
        is_active=True, verified=True, created_at=datetime.now(timezone.utc).isoformat()
    )
    db.add(new_user)
    req.status = "approved"
    db.commit()
    return {"message": "تمت الموافقة على الحساب وتفعيله بنجاح"}

@router.post("/registration-requests/{request_id}/reject")
def reject_registration(request_id: str, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    req = db.query(RegistrationRequest).filter(RegistrationRequest.id == request_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    req.status = "rejected"
    db.commit()
    return {"message": "تم رفض طلب التسجيل"}


@router.post("/bulk-import/pharmacies")
async def bulk_import_pharmacies(file: UploadFile = File(...), current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    import openpyxl, io
    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []
    if filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        rows = list(reader)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    added, errors = 0, []
    for idx, data in enumerate(rows, start=2):
        email = str(data.get("email", "") or "").strip()
        name = str(data.get("name", "") or "").strip()
        if not email or not name:
            errors.append({"row": idx, "error": "name and email required"})
            continue
        if db.query(User).filter(User.email == email).first():
            errors.append({"row": idx, "error": f"duplicate email {email}"})
            continue
        new_user = User(
            id=f"ph_{uuid.uuid4().hex[:8]}", role="pharmacy", name=name,
            email=email, password=hash_password(str(data.get("password", "123456"))),
            phone=str(data.get("phone", "") or ""), city=str(data.get("city", "") or ""),
            address=str(data.get("address", "") or ""), province=str(data.get("province", "") or ""),
            district=str(data.get("district", "") or ""), area=str(data.get("area", "") or ""),
            lat=float(data.get("lat")) if data.get("lat") else None,
            lng=float(data.get("lng")) if data.get("lng") else None,
            open_hours=str(data.get("open_hours", "") or ""),
            verified=True, is_active=True, created_at=datetime.now(timezone.utc).isoformat(),
        )
        db.add(new_user)
        added += 1
    db.commit()
    return {"added": added, "errors": errors, "message": f"تم استيراد {added} صيدلية"}


@router.post("/bulk-import/warehouses")
async def bulk_import_warehouses(file: UploadFile = File(...), current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    import openpyxl, io
    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []
    if filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(text.splitlines())
        rows = list(reader)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    added, errors = 0, []
    for idx, data in enumerate(rows, start=2):
        email = str(data.get("email", "") or "").strip()
        name = str(data.get("name", "") or "").strip()
        if not email or not name:
            errors.append({"row": idx, "error": "name and email required"})
            continue
        if db.query(User).filter(User.email == email).first():
            errors.append({"row": idx, "error": f"duplicate email {email}"})
            continue
        new_user = User(
            id=f"wh_{uuid.uuid4().hex[:8]}", role="warehouse", name=name,
            email=email, password=hash_password(str(data.get("password", "123456"))),
            phone=str(data.get("phone", "") or ""), city=str(data.get("city", "") or ""),
            address=str(data.get("address", "") or ""), open_hours=str(data.get("open_hours", "") or ""),
            verified=True, is_active=True, created_at=datetime.now(timezone.utc).isoformat(),
        )
        db.add(new_user)
        added += 1
    db.commit()
    return {"added": added, "errors": errors, "message": f"تم استيراد {added} مستودع"}

