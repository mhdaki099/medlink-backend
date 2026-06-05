from fastapi import APIRouter, Query, HTTPException, Depends, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import distinct
import uuid
from datetime import datetime, timezone

from db import get_db
from models import User, Medicine, FavoriteMedicine, CartItem, Favorite, Order, Review
from auth_utils import get_current_user, require_role
from utils.helpers import model_to_dict, safe_update

router = APIRouter()

@router.get("")
def list_pharmacies(
    lat: float = Query(None),
    lng: float = Query(None),
    radius_km: float = Query(4.0),
    province: str = Query(None),
    district: str = Query(None),
    area: str = Query(None),
    db: Session = Depends(get_db),
):
    pharmacies = db.query(User).filter(User.role == "pharmacy", User.is_active == True).all()
    results = []
    for p in pharmacies:
        if province and (p.province or "").lower() != province.lower():
            continue
        if district and (p.district or "").lower() != district.lower():
            continue
        if area and (p.area or "").lower() != area.lower():
            continue
        pd = model_to_dict(p, ["password"])
        if lat is not None and lng is not None and p.lat and p.lng:
            from math import radians, sin, cos, sqrt, atan2
            r = 6371
            dlat = radians(p.lat - lat)
            dlng = radians(p.lng - lng)
            a = sin(dlat / 2) ** 2 + cos(radians(lat)) * cos(radians(p.lat)) * sin(dlng / 2) ** 2
            dist = 2 * r * atan2(sqrt(a), sqrt(1 - a))
            pd["distance_km"] = round(dist, 2)
            if dist > radius_km:
                continue
        results.append(pd)
    if lat is not None and lng is not None:
        results.sort(key=lambda x: x.get("distance_km", 9999))
    return results

# --- Static Medicine Routes FIRST ---

@router.get("/medicines/categories")
def get_categories(db: Session = Depends(get_db)):
    cats = db.query(distinct(Medicine.category)).all()
    return [c[0] for c in cats if c[0]]

@router.get("/medicines/all")
def all_medicines(category: str = Query(None), db: Session = Depends(get_db)):
    query = db.query(Medicine, User.name.label("pharmacy_name")).join(User, Medicine.pharmacy_id == User.id)
    if category:
        query = query.filter(Medicine.category == category)
    results = []
    for med, p_name in query.all():
        m_dict = model_to_dict(med)
        if m_dict:
            m_dict["pharmacy_name"] = p_name
            results.append(m_dict)
    return results

@router.get("/medicines/favorites/{patient_id}")
def get_favorite_medicines(patient_id: str, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    favs = db.query(Medicine, User.name.label("pharmacy_name")).join(FavoriteMedicine, Medicine.id == FavoriteMedicine.medicine_id).join(User, Medicine.pharmacy_id == User.id).filter(FavoriteMedicine.user_id == patient_id).all()
    results = []
    for med, p_name in favs:
        m_dict = model_to_dict(med)
        if m_dict:
            m_dict["pharmacy_name"] = p_name
            m_dict["is_favorite"] = True
            results.append(m_dict)
    return results

# --- Cart Routes ---

@router.get("/medicines/cart/{patient_id}")
def get_cart_items(patient_id: str, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    cart = db.query(CartItem, Medicine, User.name.label("pharmacy_name")).join(Medicine, CartItem.medicine_id == Medicine.id).join(User, Medicine.pharmacy_id == User.id).filter(CartItem.user_id == patient_id).all()
    results = []
    for item, med, p_name in cart:
        m_dict = model_to_dict(med)
        if m_dict:
            m_dict["pharmacy_name"] = p_name
            m_dict["quantity"] = item.quantity
            m_dict["cart_item_id"] = item.id
            results.append(m_dict)
    return results

@router.post("/medicines/cart/add/{medicine_id}")
def add_to_cart(medicine_id: str, patient_id: str = Query(...), current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(CartItem).filter(CartItem.user_id == patient_id, CartItem.medicine_id == medicine_id).first()
    if item:
        item.quantity += 1
    else:
        item = CartItem(id=f"cart_{uuid.uuid4().hex[:8]}", user_id=patient_id, medicine_id=medicine_id, quantity=1, created_at=datetime.now(timezone.utc).isoformat())
        db.add(item)
    db.commit(); db.refresh(item)
    return {"message": "تمت الإضافة", "quantity": item.quantity}

@router.post("/medicines/cart/decrease/{medicine_id}")
def decrease_cart_item(medicine_id: str, patient_id: str = Query(...), current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(CartItem).filter(CartItem.user_id == patient_id, CartItem.medicine_id == medicine_id).first()
    if item:
        if item.quantity > 1:
            item.quantity -= 1
            db.commit()
            return {"message": "تم التنقيص", "quantity": item.quantity}
        else:
            db.delete(item); db.commit()
            return {"message": "تم الحذف", "quantity": 0}
    return {"message": "غير موجود", "quantity": 0}

@router.delete("/medicines/cart/remove/{medicine_id}")
def remove_from_cart(medicine_id: str, patient_id: str = Query(...), current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(CartItem).filter(CartItem.user_id == patient_id, CartItem.medicine_id == medicine_id).first()
    if item:
        db.delete(item); db.commit()
    return {"message": "تم الحذف"}

# --- Medicine detail routes (before dynamic pharmacy route) ---

@router.get("/medicines/{medicine_id}/details")
def get_medicine_details(medicine_id: str, db: Session = Depends(get_db)):
    med = db.query(Medicine).filter(Medicine.id == medicine_id).first()
    if not med:
        raise HTTPException(404, "الدواء غير موجود")
    d = model_to_dict(med)
    pharmacy = db.query(User).filter(User.id == med.pharmacy_id).first()
    if pharmacy:
        d["pharmacy"] = model_to_dict(pharmacy, ["password"])
    return d

@router.post("/medicines/{medicine_id}/favorite")
def toggle_medicine_favorite(medicine_id: str, patient_id: str = Query(...), current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    fav = db.query(FavoriteMedicine).filter(FavoriteMedicine.user_id == patient_id, FavoriteMedicine.medicine_id == medicine_id).first()
    if fav:
        db.delete(fav); db.commit()
        return {"is_favorite": False}
    db.add(FavoriteMedicine(id=f"favm_{uuid.uuid4().hex[:8]}", user_id=patient_id, medicine_id=medicine_id, created_at=datetime.now(timezone.utc).isoformat()))
    db.commit()
    return {"is_favorite": True}

@router.get("/medicines/{medicine_id}")
def get_medicine(medicine_id: str, db: Session = Depends(get_db)):
    med = db.query(Medicine).filter(Medicine.id == medicine_id).first()
    if not med:
        raise HTTPException(404, "الدواء غير موجود")
    return model_to_dict(med)

@router.put("/medicines/{medicine_id}")
def update_medicine(medicine_id: str, updates: dict, current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    med = db.query(Medicine).filter(Medicine.id == medicine_id).first()
    if not med:
        raise HTTPException(404, "الدواء غير موجود")
    if current_user.get("role") == "pharmacy" and current_user.get("sub") != med.pharmacy_id:
        raise HTTPException(403, "غير مصرح لك بتعديل هذا الدواء")
    for key, value in updates.items():
        if hasattr(med, key):
            setattr(med, key, value)
    db.commit(); db.refresh(med)
    return model_to_dict(med)

@router.delete("/medicines/{medicine_id}")
def delete_medicine(medicine_id: str, current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    med = db.query(Medicine).filter(Medicine.id == medicine_id).first()
    if current_user.get("role") == "pharmacy" and med and current_user.get("sub") != med.pharmacy_id:
        raise HTTPException(403, "غير مصرح لك بحذف هذا الدواء")
    if med:
        db.delete(med); db.commit()
    return {"message": "تم الحذف"}

@router.post("/medicines")
def add_medicine(medicine: dict, current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    pharmacy_id = medicine.get("pharmacy_id")
    if current_user.get("role") == "pharmacy":
        pharmacy_id = current_user.get("sub")
    if not pharmacy_id:
        raise HTTPException(400, "معرف الصيدلية مطلوب")
    med_id = f"m_{uuid.uuid4().hex[:8]}"
    new_med = Medicine(
        id=med_id, pharmacy_id=pharmacy_id, name=medicine.get("name"),
        name_en=medicine.get("name_en"), category=medicine.get("category") or None,
        price=medicine.get("price"), old_price=medicine.get("old_price"),
        description=medicine.get("description"), manufacturer=medicine.get("manufacturer"),
        stock_status=medicine.get("stock_status", "in_stock"), quantity=medicine.get("quantity", 0),
        dosage=medicine.get("dosage"), strength=medicine.get("strength"), barcode=medicine.get("barcode"),
        image=medicine.get("image"), alternatives=medicine.get("alternatives", []),
        requires_prescription=medicine.get("requires_prescription", False),
        active_ingredients=medicine.get("active_ingredients"),
        usage_info=medicine.get("usage_info"),
        side_effects=medicine.get("side_effects"),
        warnings=medicine.get("warnings"),
    )
    db.add(new_med); db.commit(); db.refresh(new_med)
    return model_to_dict(new_med)

@router.post("/medicines/bulk-upload")
async def bulk_upload_medicines(pharmacy_id: str = Query(...), file: UploadFile = File(...), current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("name"):
            continue
        med = Medicine(
            id=f"m_{uuid.uuid4().hex[:8]}", pharmacy_id=pharmacy_id,
            name=str(data.get("name", "")), name_en=str(data.get("name_en", "") or ""),
            category=str(data.get("category", "") or ""), price=float(data.get("price", 0) or 0),
            description=str(data.get("description", "") or ""), manufacturer=str(data.get("manufacturer", "") or ""),
            stock_status=str(data.get("stock_status", "in_stock") or "in_stock"),
            quantity=int(data.get("quantity", 0) or 0), dosage=str(data.get("dosage", "") or ""),
            strength=str(data.get("strength", "") or ""),
            barcode=str(data.get("barcode", "") or ""),
            active_ingredients=str(data.get("active_ingredients", "") or ""),
            usage_info=str(data.get("usage_info", "") or ""),
            side_effects=str(data.get("side_effects", "") or ""),
            warnings=str(data.get("warnings", "") or ""),
        )
        db.add(med); added += 1
    db.commit()
    return {"message": f"تم إضافة {added} دواء بنجاح", "count": added}


@router.post("/medicines/upload-excel")
async def upload_excel_alias(pharmacy_id: str = Query(...), file: UploadFile = File(...), current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    return await bulk_upload_medicines(pharmacy_id, file, current_user, db)


@router.get("/{pharmacy_id}/analytics")
def get_pharmacy_analytics(pharmacy_id: str, current_user: dict = Depends(require_role("pharmacy", "admin")), db: Session = Depends(get_db)):
    favorites = db.query(Favorite).filter(Favorite.target_id == pharmacy_id).count()
    orders = db.query(Order).filter(Order.pharmacy_id == pharmacy_id).all()
    week_ago = datetime.now(timezone.utc).timestamp() - 7 * 24 * 3600
    month_ago = datetime.now(timezone.utc).timestamp() - 30 * 24 * 3600
    weekly = 0
    monthly = 0
    for order in orders:
        try:
            ts = datetime.fromisoformat(order.created_at).timestamp()
            weekly += 1 if ts >= week_ago else 0
            monthly += 1 if ts >= month_ago else 0
        except Exception:
            pass
    customers = db.query(Order.patient_id).filter(Order.pharmacy_id == pharmacy_id).distinct().count()
    reviews = db.query(Review).filter(Review.target_id == pharmacy_id).all()
    rating = round(sum(r.rating for r in reviews) / len(reviews), 1) if reviews else 0
    return {
        "favorites_count": favorites,
        "weekly_bookings": weekly,
        "monthly_bookings": monthly,
        "total_orders": len(orders),
        "active_customers": customers,
        "overall_rating": rating,
        "total_reviews": len(reviews),
    }

# --- Dynamic pharmacy routes LAST (to avoid catching /medicines/* paths) ---

@router.get("/{pharmacy_id}")
def get_pharmacy(pharmacy_id: str, db: Session = Depends(get_db)):
    p = db.query(User).filter(User.id == pharmacy_id, User.role == "pharmacy").first()
    if not p:
        raise HTTPException(404, "الصيدلية غير موجودة")
    pd = model_to_dict(p, ["password"])
    week_ago = datetime.now(timezone.utc).timestamp() - 7 * 24 * 3600
    month_ago = datetime.now(timezone.utc).timestamp() - 30 * 24 * 3600
    orders = db.query(Order).filter(Order.pharmacy_id == pharmacy_id).all()
    weekly = monthly = 0
    for order in orders:
        try:
            ts = datetime.fromisoformat(order.created_at).timestamp()
            if ts >= week_ago: weekly += 1
            if ts >= month_ago: monthly += 1
        except Exception:
            pass
    pd["favorites_count"] = db.query(Favorite).filter(Favorite.target_id == pharmacy_id).count()
    pd["weekly_bookings"] = weekly
    pd["monthly_bookings"] = monthly
    return pd

@router.get("/{pharmacy_id}/medicines")
def get_pharmacy_medicines(pharmacy_id: str, category: str = Query(None), db: Session = Depends(get_db)):
    query = db.query(Medicine, User.name.label("pharmacy_name")).join(User, Medicine.pharmacy_id == User.id).filter(Medicine.pharmacy_id == pharmacy_id)
    if category:
        query = query.filter(Medicine.category == category)
    results = []
    for med, p_name in query.all():
        m_dict = model_to_dict(med)
        if m_dict:
            m_dict["pharmacy_name"] = p_name
            results.append(m_dict)
    return results
