"""
Drug Catalog Router — centralized drug database for prescription autocomplete
and pharmacy medicine auto-fill.
"""
from fastapi import APIRouter, Query, HTTPException, Depends, UploadFile, File
from sqlalchemy.orm import Session
import uuid
from datetime import datetime, timezone

from db import get_db
from models import DrugCatalog
from auth_utils import get_current_user, require_role
from utils.helpers import model_to_dict

router = APIRouter()

STARTER_DRUGS = [
    {"trade_name": "باراسيتامول", "trade_name_en": "Paracetamol", "active_ingredient": "Paracetamol", "strength": "500mg", "dosage_form": "tablet", "category": "مسكنات", "requires_prescription": False},
    {"trade_name": "أموكسيسيلين", "trade_name_en": "Amoxicillin", "active_ingredient": "Amoxicillin", "strength": "500mg", "dosage_form": "capsule", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "إيبوبروفين", "trade_name_en": "Ibuprofen", "active_ingredient": "Ibuprofen", "strength": "400mg", "dosage_form": "tablet", "category": "مسكنات", "requires_prescription": False},
    {"trade_name": "أوميبرازول", "trade_name_en": "Omeprazole", "active_ingredient": "Omeprazole", "strength": "20mg", "dosage_form": "capsule", "category": "هضمية", "requires_prescription": False},
    {"trade_name": "ميتفورمين", "trade_name_en": "Metformin", "active_ingredient": "Metformin", "strength": "500mg", "dosage_form": "tablet", "category": "سكري", "requires_prescription": True},
    {"trade_name": "أتورفاستاتين", "trade_name_en": "Atorvastatin", "active_ingredient": "Atorvastatin", "strength": "20mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "أملوديبين", "trade_name_en": "Amlodipine", "active_ingredient": "Amlodipine", "strength": "5mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "سيتريزين", "trade_name_en": "Cetirizine", "active_ingredient": "Cetirizine", "strength": "10mg", "dosage_form": "tablet", "category": "حساسية", "requires_prescription": False},
    {"trade_name": "أزيثروميسين", "trade_name_en": "Azithromycin", "active_ingredient": "Azithromycin", "strength": "500mg", "dosage_form": "tablet", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "ديكلوفيناك", "trade_name_en": "Diclofenac", "active_ingredient": "Diclofenac", "strength": "50mg", "dosage_form": "tablet", "category": "مسكنات", "requires_prescription": False},
    {"trade_name": "لوراتادين", "trade_name_en": "Loratadine", "active_ingredient": "Loratadine", "strength": "10mg", "dosage_form": "tablet", "category": "حساسية", "requires_prescription": False},
    {"trade_name": "رانيتيدين", "trade_name_en": "Ranitidine", "active_ingredient": "Ranitidine", "strength": "150mg", "dosage_form": "tablet", "category": "هضمية", "requires_prescription": False},
    {"trade_name": "ميترونيدازول", "trade_name_en": "Metronidazole", "active_ingredient": "Metronidazole", "strength": "500mg", "dosage_form": "tablet", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "كلاريثروميسين", "trade_name_en": "Clarithromycin", "active_ingredient": "Clarithromycin", "strength": "500mg", "dosage_form": "tablet", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "فلوكونازول", "trade_name_en": "Fluconazole", "active_ingredient": "Fluconazole", "strength": "150mg", "dosage_form": "capsule", "category": "مضادات فطرية", "requires_prescription": True},
    {"trade_name": "بريدنيزولون", "trade_name_en": "Prednisolone", "active_ingredient": "Prednisolone", "strength": "5mg", "dosage_form": "tablet", "category": "كورتيزون", "requires_prescription": True},
    {"trade_name": "سالبوتامول", "trade_name_en": "Salbutamol", "active_ingredient": "Salbutamol", "strength": "100mcg", "dosage_form": "inhaler", "category": "تنفسية", "requires_prescription": True},
    {"trade_name": "ليفوثيروكسين", "trade_name_en": "Levothyroxine", "active_ingredient": "Levothyroxine", "strength": "50mcg", "dosage_form": "tablet", "category": "غدة درقية", "requires_prescription": True},
    {"trade_name": "إنالابريل", "trade_name_en": "Enalapril", "active_ingredient": "Enalapril", "strength": "5mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "فيتامين د", "trade_name_en": "Vitamin D3", "active_ingredient": "Cholecalciferol", "strength": "1000IU", "dosage_form": "tablet", "category": "فيتامينات", "requires_prescription": False},
    {"trade_name": "كالسيوم", "trade_name_en": "Calcium Carbonate", "active_ingredient": "Calcium Carbonate", "strength": "500mg", "dosage_form": "tablet", "category": "فيتامينات", "requires_prescription": False},
    {"trade_name": "أسبرين", "trade_name_en": "Aspirin", "active_ingredient": "Acetylsalicylic acid", "strength": "100mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": False},
    {"trade_name": "كلوبيدوغريل", "trade_name_en": "Clopidogrel", "active_ingredient": "Clopidogrel", "strength": "75mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "سيمفاستاتين", "trade_name_en": "Simvastatin", "active_ingredient": "Simvastatin", "strength": "20mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "ليزينوبريل", "trade_name_en": "Lisinopril", "active_ingredient": "Lisinopril", "strength": "10mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "ميتوبرولول", "trade_name_en": "Metoprolol", "active_ingredient": "Metoprolol", "strength": "50mg", "dosage_form": "tablet", "category": "قلبية", "requires_prescription": True},
    {"trade_name": "فوروسيميد", "trade_name_en": "Furosemide", "active_ingredient": "Furosemide", "strength": "40mg", "dosage_form": "tablet", "category": "مدرات البول", "requires_prescription": True},
    {"trade_name": "سيبروفلوكساسين", "trade_name_en": "Ciprofloxacin", "active_ingredient": "Ciprofloxacin", "strength": "500mg", "dosage_form": "tablet", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "دوكسيسيكلين", "trade_name_en": "Doxycycline", "active_ingredient": "Doxycycline", "strength": "100mg", "dosage_form": "capsule", "category": "مضادات حيوية", "requires_prescription": True},
    {"trade_name": "ترامادول", "trade_name_en": "Tramadol", "active_ingredient": "Tramadol", "strength": "50mg", "dosage_form": "capsule", "category": "مسكنات", "requires_prescription": True},
]


def seed_catalog(db: Session):
    """Seed starter drugs if catalog is empty."""
    if db.query(DrugCatalog).count() > 0:
        return
    for d in STARTER_DRUGS:
        db.add(DrugCatalog(
            id=f"dc_{uuid.uuid4().hex[:8]}",
            trade_name=d["trade_name"],
            trade_name_en=d.get("trade_name_en"),
            active_ingredient=d.get("active_ingredient"),
            strength=d.get("strength"),
            dosage_form=d.get("dosage_form"),
            category=d.get("category"),
            requires_prescription=d.get("requires_prescription", False),
        ))
    db.commit()


@router.get("/search")
def search_drugs(q: str = Query(""), db: Session = Depends(get_db)):
    """Search drug catalog by trade name or active ingredient."""
    seed_catalog(db)
    if not q:
        return db.query(DrugCatalog).limit(50).all().__class__([model_to_dict(d) for d in db.query(DrugCatalog).limit(50).all()])
    results = db.query(DrugCatalog).filter(
        DrugCatalog.trade_name.ilike(f"%{q}%") |
        DrugCatalog.trade_name_en.ilike(f"%{q}%") |
        DrugCatalog.active_ingredient.ilike(f"%{q}%")
    ).limit(20).all()
    return [model_to_dict(d) for d in results]


@router.get("")
def list_drugs(category: str = Query(None), db: Session = Depends(get_db)):
    seed_catalog(db)
    query = db.query(DrugCatalog)
    if category:
        query = query.filter(DrugCatalog.category == category)
    return [model_to_dict(d) for d in query.all()]


@router.get("/categories")
def get_drug_categories(db: Session = Depends(get_db)):
    from sqlalchemy import distinct
    seed_catalog(db)
    cats = db.query(distinct(DrugCatalog.category)).all()
    return [c[0] for c in cats if c[0]]


@router.get("/{drug_id}")
def get_drug(drug_id: str, db: Session = Depends(get_db)):
    d = db.query(DrugCatalog).filter(DrugCatalog.id == drug_id).first()
    if not d:
        raise HTTPException(404, "الدواء غير موجود")
    return model_to_dict(d)


@router.post("")
def add_drug(data: dict, current_user: dict = Depends(require_role("admin")), db: Session = Depends(get_db)):
    drug = DrugCatalog(
        id=f"dc_{uuid.uuid4().hex[:8]}",
        trade_name=data.get("trade_name", ""),
        trade_name_en=data.get("trade_name_en"),
        active_ingredient=data.get("active_ingredient"),
        strength=data.get("strength"),
        dosage_form=data.get("dosage_form"),
        manufacturer=data.get("manufacturer"),
        barcode=data.get("barcode"),
        category=data.get("category"),
        usage_info=data.get("usage_info"),
        side_effects=data.get("side_effects"),
        warnings=data.get("warnings"),
        requires_prescription=data.get("requires_prescription", False),
    )
    db.add(drug)
    db.commit()
    db.refresh(drug)
    return model_to_dict(drug)


@router.post("/bulk-upload")
async def bulk_upload_drugs(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role("admin")),
    db: Session = Depends(get_db)
):
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in ws[1]]
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("trade_name"):
            continue
        drug = DrugCatalog(
            id=f"dc_{uuid.uuid4().hex[:8]}",
            trade_name=str(data.get("trade_name", "")),
            trade_name_en=str(data.get("trade_name_en", "") or ""),
            active_ingredient=str(data.get("active_ingredient", "") or ""),
            strength=str(data.get("strength", "") or ""),
            dosage_form=str(data.get("dosage_form", "") or ""),
            manufacturer=str(data.get("manufacturer", "") or ""),
            barcode=str(data.get("barcode", "") or ""),
            category=str(data.get("category", "") or ""),
            usage_info=str(data.get("usage_info", "") or ""),
            side_effects=str(data.get("side_effects", "") or ""),
            warnings=str(data.get("warnings", "") or ""),
            requires_prescription=bool(data.get("requires_prescription", False)),
        )
        db.add(drug)
        added += 1
    db.commit()
    return {"message": f"تم إضافة {added} دواء بنجاح", "count": added}
