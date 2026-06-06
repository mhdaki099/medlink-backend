from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from sqlalchemy.orm import Session
from datetime import datetime, timezone
import uuid

from db import get_db
from models import User, WarehouseInventory, WarehouseOrder, Notification, WarehousePromoter
from auth_utils import get_current_user, require_role
from utils.helpers import model_to_dict

router = APIRouter()


def _order_item_qty(item: dict) -> int:
    return max(1, int(item.get("qty") or item.get("quantity") or 1))


def _deduct_warehouse_inventory(db: Session, raw_items: list) -> None:
    enriched = _enrich_warehouse_order_items(db, raw_items)
    for item in enriched:
        item_id = item.get("item_id")
        if not item_id:
            continue
        inv = db.query(WarehouseInventory).filter(WarehouseInventory.id == item_id).first()
        if not inv:
            continue
        qty = _order_item_qty(item)
        available = inv.stock or 0
        if available < qty:
            raise HTTPException(400, f"مخزون غير كافٍ للصنف {inv.name} (متوفر: {available})")
        inv.stock = available - qty


def _restore_warehouse_inventory(db: Session, raw_items: list) -> None:
    enriched = _enrich_warehouse_order_items(db, raw_items)
    for item in enriched:
        item_id = item.get("item_id")
        if not item_id:
            continue
        inv = db.query(WarehouseInventory).filter(WarehouseInventory.id == item_id).first()
        if not inv:
            continue
        inv.stock = (inv.stock or 0) + _order_item_qty(item)


def _enrich_warehouse_order_items(db: Session, raw_items: list) -> list:
    enriched = []
    for item in raw_items or []:
        row = dict(item)
        inv = None
        if row.get("item_id"):
            inv = db.query(WarehouseInventory).filter(WarehouseInventory.id == row["item_id"]).first()
        if inv:
            if not row.get("name"):
                row["name"] = inv.name
            row["unit"] = inv.unit
            if row.get("bulk_price") is None:
                row["bulk_price"] = inv.bulk_price
        enriched.append(row)
    return enriched


@router.get("")
def list_warehouses(db: Session = Depends(get_db)):
    warehouses = db.query(User).filter(User.role == "warehouse").all()
    return [model_to_dict(w, ["password"]) for w in warehouses]


@router.get("/promoters")
def list_warehouse_promoters(current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    warehouse_id = current_user.get("sub")
    if current_user.get("role") != "warehouse":
        raise HTTPException(403, "للمستودع فقط")
    rows = db.query(WarehousePromoter).filter(
        WarehousePromoter.warehouse_id == warehouse_id,
    ).order_by(WarehousePromoter.name).all()
    return [model_to_dict(r) for r in rows]


@router.get("/promoter-commissions")
def get_promoter_commissions(
    year: int = Query(None),
    month: int = Query(None),
    promoter_id: str = Query(None),
    current_user: dict = Depends(require_role("warehouse", "admin")),
    db: Session = Depends(get_db),
):
    warehouse_id = current_user.get("sub")
    if current_user.get("role") != "warehouse":
        raise HTTPException(403, "للمستودع فقط")
    now = datetime.now(timezone.utc)
    y = int(year or now.year)
    m = int(month or now.month)
    if m < 1 or m > 12:
        raise HTTPException(400, "الشهر غير صالح")
    month_key = f"{y:04d}-{m:02d}"

    promoter_rows = {
        p.id: model_to_dict(p)
        for p in db.query(WarehousePromoter).filter(WarehousePromoter.warehouse_id == warehouse_id).all()
    }

    orders = db.query(WarehouseOrder).filter(
        WarehouseOrder.warehouse_id == warehouse_id,
        WarehouseOrder.status.in_(("shipped", "delivered")),
    ).all()

    by_promoter: dict = {}
    for order in orders:
        inv = order.invoice if isinstance(order.invoice, dict) else {}
        promo = inv.get("promoter") or {}
        pid = promo.get("id")
        if not pid and not promo.get("name"):
            continue
        if promoter_id and pid != promoter_id:
            continue

        inv_date = (inv.get("date") or order.created_at or "")[:7]
        if inv_date != month_key:
            continue

        total_inv = float(inv.get("total") or order.total or 0)
        pct = float(promo.get("commission_percent") or 0)
        amount = float(promo.get("commission_amount") or 0)
        if not amount and pct:
            amount = round(total_inv * pct / 100, 2)

        ph = db.query(User).filter(User.id == order.pharmacy_id).first()
        order_row = {
            "order_id": order.id,
            "purchase_order_number": order.purchase_order_number,
            "invoice_number": inv.get("number"),
            "invoice_date": inv.get("date"),
            "pharmacy_name": ph.name if ph else "صيدلية",
            "invoice_total": round(total_inv, 2),
            "commission_percent": pct,
            "commission_amount": round(amount, 2),
            "order_status": order.status,
        }

        key = pid or f"name:{promo.get('name')}"
        if key not in by_promoter:
            db_promo = promoter_rows.get(pid) if pid else None
            by_promoter[key] = {
                "promoter_id": pid,
                "promoter_name": promo.get("name") or (db_promo or {}).get("name", "مندوب"),
                "default_percent": (db_promo or {}).get("commission_percent"),
                "orders_count": 0,
                "total_sales": 0.0,
                "total_commission": 0.0,
                "orders": [],
            }
        bucket = by_promoter[key]
        bucket["orders_count"] += 1
        bucket["total_sales"] = round(bucket["total_sales"] + total_inv, 2)
        bucket["total_commission"] = round(bucket["total_commission"] + amount, 2)
        bucket["orders"].append(order_row)

    for pid, prow in promoter_rows.items():
        if pid in by_promoter or (promoter_id and promoter_id != pid):
            continue
        by_promoter[pid] = {
            "promoter_id": pid,
            "promoter_name": prow.get("name"),
            "default_percent": prow.get("commission_percent"),
            "orders_count": 0,
            "total_sales": 0.0,
            "total_commission": 0.0,
            "orders": [],
        }

    summary = sorted(by_promoter.values(), key=lambda x: x["total_commission"], reverse=True)
    if promoter_id:
        summary = [s for s in summary if s.get("promoter_id") == promoter_id]

    return {
        "month": month_key,
        "year": y,
        "month_num": m,
        "summary": summary,
        "grand_total_commission": round(sum(s["total_commission"] for s in summary), 2),
        "grand_total_sales": round(sum(s["total_sales"] for s in summary), 2),
        "orders_with_promoter": sum(s["orders_count"] for s in summary),
    }


@router.post("/promoters")
def create_warehouse_promoter(body: dict, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    warehouse_id = current_user.get("sub")
    if current_user.get("role") != "warehouse":
        raise HTTPException(403, "للمستودع فقط")
    name = str(body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "اسم المندوب مطلوب")
    pct = float(body.get("commission_percent") or 0)
    if pct < 0 or pct > 100:
        raise HTTPException(400, "نسبة العمولة بين 0 و 100")
    now = datetime.now(timezone.utc).isoformat()
    row = WarehousePromoter(
        id=f"wp_{uuid.uuid4().hex[:8]}",
        warehouse_id=warehouse_id,
        name=name,
        phone=str(body.get("phone") or "").strip() or None,
        commission_percent=pct,
        is_active=True,
        created_at=now,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return model_to_dict(row)


@router.put("/promoters/{promoter_id}")
def update_warehouse_promoter(promoter_id: str, body: dict, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    row = db.query(WarehousePromoter).filter(WarehousePromoter.id == promoter_id).first()
    if not row:
        raise HTTPException(404, "المندوب غير موجود")
    if current_user.get("role") == "warehouse" and current_user.get("sub") != row.warehouse_id:
        raise HTTPException(403, "غير مصرح")
    if "name" in body:
        name = str(body.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "اسم المندوب مطلوب")
        row.name = name
    if "phone" in body:
        row.phone = str(body.get("phone") or "").strip() or None
    if "commission_percent" in body:
        pct = float(body.get("commission_percent") or 0)
        if pct < 0 or pct > 100:
            raise HTTPException(400, "نسبة العمولة بين 0 و 100")
        row.commission_percent = pct
    if "is_active" in body:
        row.is_active = bool(body.get("is_active"))
    db.commit()
    db.refresh(row)
    return model_to_dict(row)


@router.delete("/promoters/{promoter_id}")
def delete_warehouse_promoter(promoter_id: str, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    row = db.query(WarehousePromoter).filter(WarehousePromoter.id == promoter_id).first()
    if not row:
        raise HTTPException(404, "المندوب غير موجود")
    if current_user.get("role") == "warehouse" and current_user.get("sub") != row.warehouse_id:
        raise HTTPException(403, "غير مصرح")
    row.is_active = False
    db.commit()
    return {"ok": True}


@router.put("/orders/{order_id}/status")
def update_order_status(order_id: str, status_update: dict, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    order = db.query(WarehouseOrder).filter(WarehouseOrder.id == order_id).first()
    if not order:
        raise HTTPException(404, "الطلب غير موجود")
    if current_user.get("role") == "warehouse" and current_user.get("sub") != order.warehouse_id:
        raise HTTPException(403, "غير مصرح لك بتحديث هذا الطلب")
    new_status = status_update.get("status", order.status)
    if new_status == "delivered":
        raise HTTPException(403, "تأكيد الاستلام يتم من الصيدلية فقط")
    allowed = {"processing", "shipped", "cancelled"}
    if new_status not in allowed:
        raise HTTPException(400, "حالة غير صالحة للمستودع")
    if new_status == "processing" and order.status not in ("pending",):
        raise HTTPException(400, "لا يمكن قبول هذا الطلب")
    if new_status == "shipped" and order.status not in ("processing",):
        raise HTTPException(400, "يجب تجهيز الطلب قبل الشحن")
    if new_status == "cancelled" and order.status not in ("pending", "processing"):
        raise HTTPException(400, "لا يمكن إلغاء هذا الطلب")
    now = datetime.now(timezone.utc).isoformat()
    old_status = order.status
    if new_status == "shipped" and old_status == "processing":
        _deduct_warehouse_inventory(db, order.items or [])
    if new_status == "cancelled" and old_status == "processing":
        _restore_warehouse_inventory(db, order.items or [])
    order.status = new_status
    if "delivery_time" in status_update:
        order.delivery_time = status_update["delivery_time"]
    if new_status == "shipped":
        from routers.orders import ensure_warehouse_invoice
        ensure_warehouse_invoice(db, order)
        pharmacy = db.query(User).filter(User.id == order.pharmacy_id).first()
        db.add(Notification(
            id=f"ntf_{uuid.uuid4().hex[:8]}",
            user_id=order.pharmacy_id,
            title="شحنة من المستودع 🚚",
            message=f"تم شحن أمر الشراء {order.purchase_order_number or order.id} — يرجى تأكيد الاستلام عند الوصول",
            type="warehouse_order",
            created_at=now,
        ))
    db.commit(); db.refresh(order)
    return model_to_dict(order)

@router.post("/orders")
def create_order(order: dict, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    from routers.orders import _generate_purchase_order_number
    wo_id = f"wo_{uuid.uuid4().hex[:8]}"
    now = datetime.now(timezone.utc).isoformat()
    po_number = _generate_purchase_order_number(wo_id, now)
    new_wo = WarehouseOrder(
        id=wo_id, pharmacy_id=order.get("pharmacy_id"), warehouse_id=order.get("warehouse_id"),
        purchase_order_number=po_number,
        items=order.get("items", []), total=order.get("total", 0), status="pending", created_at=now,
    )
    db.add(new_wo); db.commit(); db.refresh(new_wo)
    return model_to_dict(new_wo)

@router.get("/{warehouse_id}")
def get_warehouse(warehouse_id: str, db: Session = Depends(get_db)):
    w = db.query(User).filter(User.id == warehouse_id, User.role == "warehouse").first()
    if not w:
        raise HTTPException(404, "المستودع غير موجود")
    return model_to_dict(w, ["password"])

@router.get("/{warehouse_id}/inventory")
def get_inventory(warehouse_id: str, db: Session = Depends(get_db)):
    items = db.query(WarehouseInventory).filter(WarehouseInventory.warehouse_id == warehouse_id).all()
    return [model_to_dict(i) for i in items]


@router.post("/{warehouse_id}/inventory")
def add_inventory_item(warehouse_id: str, item: dict, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    if current_user.get("role") == "warehouse" and current_user.get("sub") != warehouse_id:
        raise HTTPException(403, "غير مصرح لك بإضافة مخزون لهذا المستودع")
    new_item = WarehouseInventory(
        id=f"wi_{uuid.uuid4().hex[:8]}",
        warehouse_id=warehouse_id,
        name=item.get("name", ""),
        category=item.get("category") or None,
        strength=item.get("strength"),
        barcode=item.get("barcode"),
        bulk_price=float(item.get("bulk_price", item.get("price", 0)) or 0),
        unit=item.get("unit", "علبة"),
        stock=int(item.get("stock", item.get("quantity", 0)) or 0),
        min_order=int(item.get("min_order", 1) or 1),
    )
    if not new_item.name:
        raise HTTPException(400, "اسم الصنف مطلوب")
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return model_to_dict(new_item)


@router.put("/inventory/{item_id}")
def update_inventory_item(item_id: str, updates: dict, current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    item = db.query(WarehouseInventory).filter(WarehouseInventory.id == item_id).first()
    if not item:
        raise HTTPException(404, "الصنف غير موجود")
    if current_user.get("role") == "warehouse" and current_user.get("sub") != item.warehouse_id:
        raise HTTPException(403, "غير مصرح")
    if "stock_add" in updates:
        add_qty = max(0, int(updates.get("stock_add") or 0))
        item.stock = (item.stock or 0) + add_qty
    for key in ["name", "category", "strength", "barcode", "bulk_price", "unit", "stock", "min_order"]:
        if key in updates and key != "stock_add":
            setattr(item, key, updates[key])
    db.commit()
    db.refresh(item)
    result = model_to_dict(item)
    if updates.get("invoice_number"):
        result["invoice_number"] = updates.get("invoice_number")
    return result

@router.get("/{warehouse_id}/orders")
def get_orders(warehouse_id: str, current_user: dict = Depends(require_role("warehouse", "pharmacy", "admin")), db: Session = Depends(get_db)):
    orders = db.query(WarehouseOrder).filter(WarehouseOrder.warehouse_id == warehouse_id).all()
    results = []
    for o in orders:
        odict = model_to_dict(o)
        odict["items"] = _enrich_warehouse_order_items(db, odict.get("items") or [])
        ph = db.query(User).filter(User.id == o.pharmacy_id).first()
        if ph:
            odict["pharmacy"] = model_to_dict(ph, ["password"])
        results.append(odict)
    return results

@router.post("/{warehouse_id}/bulk-upload")
async def bulk_upload_inventory(warehouse_id: str, file: UploadFile = File(...), current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("name"):
            continue
        item = WarehouseInventory(
            id=f"wi_{uuid.uuid4().hex[:8]}", warehouse_id=warehouse_id,
            name=str(data.get("name", "")), category=str(data.get("category", "") or ""),
            strength=str(data.get("strength", "") or ""),
            barcode=str(data.get("barcode", "") or ""),
            bulk_price=float(data.get("bulk_price", 0) or data.get("price", 0) or 0),  # FIX: use bulk_price column
            stock=int(data.get("stock", 0) or data.get("quantity", 0) or 0),
            unit=str(data.get("unit", "") or ""),
            min_order=int(data.get("min_order", 1) or 1),
        )
        db.add(item); added += 1
    db.commit()
    return {"message": f"تم إضافة {added} عنصر بنجاح", "count": added}


@router.post("/{warehouse_id}/inventory/upload-excel")
async def upload_inventory_excel_alias(warehouse_id: str, file: UploadFile = File(...), current_user: dict = Depends(require_role("warehouse", "admin")), db: Session = Depends(get_db)):
    return await bulk_upload_inventory(warehouse_id, file, current_user, db)
